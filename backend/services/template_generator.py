import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from config import Config

class TemplateGenerator:
    
    @staticmethod
    def generate_walkin_template():
        """Generate walk-in data upload template"""
        template_data = {
            'Mall Name': ['Gopalan Signature Tower'],
            'Date': [datetime.now().strftime('%Y-%m-%d')],
            'Footfall': [5000],
            'Peak Hour Visitors': [1200],
            'Peak Hour Start': ['18:00'],
            'Peak Hour End': ['20:00'],
            'Average Dwell Time': [45],
            'Weather Condition': ['Sunny'],
            'Special Event': ['Weekend Sale']
        }
        
        df = pd.DataFrame(template_data)
        return TemplateGenerator._style_excel_template(df, 'Walk-in Data Template')
    
    @staticmethod
    def generate_sales_template():
        """Generate sales data upload template"""
        template_data = {
            'Mall Name': ['Gopalan Signature Tower'],
            'Brand Name': ['Nike'],
            'Date': [datetime.now().strftime('%Y-%m-%d')],
            'Total Sales': [150000],
            'Transaction Count': [75],
            'Customer Count': [68],
            'Returns Amount': [2500],
            'Discount Amount': [5000],
            'Net Sales': [142500],
            'Tax Amount': [22800]
        }
        
        df = pd.DataFrame(template_data)
        return TemplateGenerator._style_excel_template(df, 'Sales Data Template')
    
    @staticmethod
    def generate_rent_template():
        """Generate rent data upload template"""
        template_data = {
            'Mall Name': ['Gopalan Signature Tower'],
            'Brand Name': ['Nike'],
            'Month': [datetime.now().strftime('%Y-%m')],
            'Base Rent': [250000],
            'Revenue Share %': [5.0],
            'Revenue Share Amount': [7500],
            'Maintenance Charges': [25000],
            'Other Charges': [5000],
            'Total Rent': [287500],
            'Payment Status': ['Pending'],
            'Payment Date': [''],
            'Invoice Number': ['INV-2024-001']
        }
        
        df = pd.DataFrame(template_data)
        return TemplateGenerator._style_excel_template(df, 'Rent Data Template')
    
    @staticmethod
    def generate_bulk_brand_template():
        """Generate bulk brand onboarding template"""
        template_data = {
            'Mall Name': ['Gopalan Signature Tower'],
            'Brand Name': ['New Brand 1'],
            'Category': ['Fashion'],
            'Sub Category': ['Apparel'],
            'Store Area (sq ft)': [1500],
            'Lease Start Date': [datetime.now().strftime('%Y-%m-%d')],
            'Lease End Date': [(datetime.now().replace(year=datetime.now().year + 5)).strftime('%Y-%m-%d')],
            'Monthly Rent': [200000],
            'Revenue Share %': [5.0],
            'Contact Person': ['John Doe'],
            'Contact Email': ['john@brand.com'],
            'Contact Phone': ['9876543210']
        }
        
        df = pd.DataFrame(template_data)
        return TemplateGenerator._style_excel_template(df, 'Bulk Brand Onboarding Template')
    
    @staticmethod
    def _style_excel_template(df, title):
        """Apply styling to Excel template"""
        output_path = os.path.join(Config.UPLOAD_FOLDER, f"template_{title.replace(' ', '_')}.xlsx")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Template', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Template']
            
            # Define styles
            header_font = Font(name='Rubik', bold=True, size=11, color='FFFFFF')
            header_fill = PatternFill(start_color='4e73df', end_color='4e73df', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            cell_font = Font(name='Times New Roman', size=10)
            cell_alignment = Alignment(horizontal='left', vertical='center')
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Style header row
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                
                # Adjust column width
                column_letter = get_column_letter(col)
                worksheet.column_dimensions[column_letter].width = 20
            
            # Style data rows
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = cell_font
                    cell.alignment = cell_alignment
                    cell.border = border
            
            # Add instructions sheet
            instructions_sheet = workbook.create_sheet('Instructions')
            instructions = [
                ['INSTRUCTIONS FOR DATA UPLOAD'],
                [''],
                ['1. Do not change the column headers'],
                ['2. Fill in your data starting from row 2'],
                ['3. Required fields are marked with *'],
                ['4. Date format should be YYYY-MM-DD'],
                ['5. Month format should be YYYY-MM'],
                ['6. Ensure Mall Name and Brand Name are already onboarded'],
                ['7. Use positive numbers only for amounts and counts'],
                [''],
                ['REQUIRED FIELDS:'],
            ]
            
            # Add required fields based on template type
            if 'Walk-in' in title:
                instructions.append(['- Mall Name, Date, Footfall'])
            elif 'Sales' in title:
                instructions.append(['- Mall Name, Brand Name, Date, Total Sales, Transaction Count'])
            elif 'Rent' in title:
                instructions.append(['- Mall Name, Brand Name, Month, Base Rent, Total Rent'])
            elif 'Brand' in title:
                instructions.append(['- Mall Name, Brand Name, Category, Monthly Rent'])
            
            for row_data in instructions:
                instructions_sheet.append(row_data)
            
            # Style instructions sheet
            for row in instructions_sheet.iter_rows():
                for cell in row:
                    if cell.row == 1:
                        cell.font = Font(name='Rubik', bold=True, size=12, color='4e73df')
                    else:
                        cell.font = Font(name='Times New Roman', size=10)
            
            instructions_sheet.column_dimensions['A'].width = 80
        
        return output_path
    
    @staticmethod
    def generate_all_templates():
        """Generate all template files"""
        templates = {
            'walkin': TemplateGenerator.generate_walkin_template(),
            'sales': TemplateGenerator.generate_sales_template(),
            'rent': TemplateGenerator.generate_rent_template(),
            'brands': TemplateGenerator.generate_bulk_brand_template()
        }
        return templates